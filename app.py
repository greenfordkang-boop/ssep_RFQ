import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import traceback
import json
import os
from datetime import datetime

# =========================================================
# [핵심] 안전하게 엑셀에 값을 넣는 함수
# =========================================================
def safe_write(ws, row, col, value):
    try:
        cell = ws.cell(row=row, column=col)
        cell.value = value
    except AttributeError:
        pass  # 병합된 셀 에러 무시
    except Exception:
        pass

# =========================================================
# [설정] 엑셀 좌표 (이미지 양식 기준)
# =========================================================
# 재료비 섹션
MAT_START_ROW = 9
MAT_MAX_ROW = 24
COL_MAT_CODE = 3      # 부품코드 (C열)
COL_MAT_NAME = 6      # 부품명 (F열)
COL_MAT_US = 4        # U/S (D열)
COL_MAT_SPEC = 5      # 재질/규격 (E열)
COL_MAT_UNIT = 7      # 단위 (G열)
COL_MAT_PRICE = 8     # 단가 (H열)
COL_MAT_NET = 9       # NET(g,mm) (I열)
COL_MAT_SCRAP = 10    # SCRAP(g,mm) (J열)
COL_MAT_INPUT = 11    # 투입중량 (K열)
COL_MAT_AMOUNT = 12   # 금액 (L열)
COL_MAT_LOSS_RATE = 13  # 자재LOSS율 (M열)
COL_MAT_LOSS_AMOUNT = 14  # LOSS금액 (N열)
COL_MAT_WASTE = 15    # 산업폐기물처리비용 (O열)
COL_MAT_DIE_LOSS = 16 # 다이캐스팅LOSS인정 (P열)
COL_MAT_DIE_AMOUNT = 17  # 금액 (Q열)
COL_MAT_TOTAL = 18    # 재료비 (R열)

# 가공비 섹션 (엑셀 템플릿 기준: (2) 가공비 표의 열 위치)
PRO_START_ROW = 27
PRO_MAX_ROW = 45
COL_PRO_NAME = 3      # C열: 부품명
COL_PRO_US = 5        # E열: U/S
COL_PRO_PROCESS = 6   # F열: 공정명
COL_PRO_MACH = 7      # G열: 사용기계
COL_PRO_MAN = 9       # I열: 인
COL_PRO_TIME = 10     # J열: 공수(SEC)
COL_PRO_RATE = 11     # K열: 임율(원/HR)
COL_PRO_AMOUNT1 = 12  # L열: 금액(원/EA)
# 아래 컬럼은 현재 템플릿에서 직접 사용하지 않으므로 필요 시 확장
COL_PRO_BASIS = 13    # (옵션) 산출근거(원/HR)
COL_PRO_AMOUNT2 = 14  # (옵션) 금액(원/EA) 비
COL_PRO_PREP = 15     # (옵션) 준비시간(분)

# =========================================================
# [저장/불러오기 유틸] 결과 저장소
# =========================================================
SAVE_FILE = "saved_results.json"

def load_saved_results():
    """저장된 산출 목록 불러오기"""
    if not os.path.exists(SAVE_FILE):
        return []
    try:
        with open(SAVE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def save_results_list(results):
    """산출 목록 전체를 파일에 저장"""
    try:
        with open(SAVE_FILE, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

st.set_page_config(page_title="원가계산서 시스템", layout="wide")

# =========================================================
# [로그인] 초기 화면: 비밀번호 입력 후 본 화면 진입
# =========================================================
APP_PASSWORD = "ssep2025"

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    # 화면 중앙 정렬용 레이아웃
    _, center, _ = st.columns([1, 2, 1])
    with center:
        st.markdown("### 🔐 SSEP 원가계산 시스템")
        st.markdown("##### 로그인 후 사용 가능합니다.")
        password = st.text_input("비밀번호", type="password")
        login_btn = st.button("로그인", use_container_width=True)

        if login_btn:
            if password == APP_PASSWORD:
                st.session_state.logged_in = True
                try:
                    st.rerun()
                except AttributeError:
                    if hasattr(st, "experimental_rerun"):
                        st.experimental_rerun()
            else:
                st.error("비밀번호가 올바르지 않습니다.")

    # 로그인 전에는 이하 내용 렌더링하지 않음
    st.stop()

st.title("📋 원가계산서 작성 시스템")

# =========================================================
# [UI 1] 기본 정보 입력
# =========================================================
with st.expander("1. 기본 정보 입력", expanded=True):
    col1, col2, col3, col4 = st.columns(4)
    p_no = col1.text_input("품번", "96240-BQ000")
    p_name = col2.text_input("품명", "ANTENA ASSY-CRASH PAD")
    car = col3.text_input("차종", "QU2i")
    company = col4.text_input("업체", "")
    labor_rate = st.number_input("적용임율 (원/HR)", value=3500, min_value=0)

st.divider()

# =========================================================
# [UI 2] 재료비 산출 (편집 가능)
# =========================================================
header_col1, header_col2 = st.columns([3, 1])
with header_col1:
    st.subheader("2. 부품별 재료비 산출")
    st.caption("재료비 정보를 입력하세요. 금액과 재료비는 자동으로 계산됩니다.")
with header_col2:
    if st.button("🆕 신규 견적 작성", use_container_width=True):
        # 재료비 테이블을 초기값으로 리셋
        st.session_state.material_df = get_default_material_df()
        # 계산 결과 및 저장된 공정 데이터 초기화
        st.session_state.pop("saved_process_df", None)
        # Streamlit 1.32+에서는 st.rerun() 사용
        try:
            st.rerun()
        except AttributeError:
            # 구버전 호환
            if hasattr(st, "experimental_rerun"):
                st.experimental_rerun()

def get_default_material_df():
    """초기 재료비 테이블 (빈 페이지용 컬럼만 정의)"""
    columns = [
        "부품명",
        "부품코드",
        "U/S",
        "재질/규격",
        "단위",
        "단가",
        "NET(g,mm)",
        "SCRAP(g,mm)",
        "자재LOSS율(%)",
        "산업폐기물처리비용",
        "다이캐스팅LOSS인정",
    ]
    return pd.DataFrame(columns=columns)


def get_default_process_df():
    """초기가공비 테이블 (빈 페이지용 컬럼만 정의)"""
    columns = [
        "부품명",
        "U/S",
        "공정명",
        "사용기계",
        "인",
        "공수(SEC)",
        "준비시간(분)",
        "산출근거(원/HR)",
        "여유율(%)",
    ]
    return pd.DataFrame(columns=columns)

# 세션 상태 초기화
if 'material_df' not in st.session_state:
    st.session_state.material_df = get_default_material_df()

# 편집 가능한 테이블 생성 (계산 컬럼 포함)
# 항상 세션에 있는 입력값(material_df)을 기준으로 매 렌더링 때마다 재계산한다.
edit_df = st.session_state.material_df.copy().reset_index(drop=True)

# 숫자형 데이터 변환 및 계산
calc_df = edit_df.copy()
calc_df['단가'] = pd.to_numeric(calc_df['단가'], errors='coerce').fillna(0)
calc_df['U/S'] = pd.to_numeric(calc_df['U/S'], errors='coerce').fillna(1)
calc_df['NET(g,mm)'] = pd.to_numeric(calc_df['NET(g,mm)'], errors='coerce').fillna(0)
calc_df['자재LOSS율(%)'] = pd.to_numeric(calc_df['자재LOSS율(%)'], errors='coerce').fillna(0)
calc_df['산업폐기물처리비용'] = pd.to_numeric(calc_df['산업폐기물처리비용'], errors='coerce').fillna(0)
calc_df['다이캐스팅LOSS인정'] = pd.to_numeric(calc_df['다이캐스팅LOSS인정'], errors='coerce').fillna(0)

# 금액 계산: 단가 × NET(g,mm) × U/S
calc_df['금액'] = calc_df['단가'] * calc_df['NET(g,mm)'] * calc_df['U/S']

# 자재LOSS 금액 계산
calc_df['LOSS금액'] = calc_df['금액'] * (calc_df['자재LOSS율(%)'] / 100)

# 재료비 계산
calc_df['재료비'] = calc_df['금액'] + calc_df['LOSS금액'] + calc_df['산업폐기물처리비용'] + calc_df['다이캐스팅LOSS인정']

# 계산 컬럼 추가 (숫자형으로 명확히 설정)
edit_df['금액'] = calc_df['금액'].astype(float)
edit_df['LOSS금액'] = calc_df['LOSS금액'].astype(float)
edit_df['재료비'] = calc_df['재료비'].astype(float)

# 데이터 편집기 (편집 가능한 테이블)
edited_mat = st.data_editor(
    edit_df,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "부품명": st.column_config.TextColumn("부품명", width="large", required=True),
        "부품코드": st.column_config.TextColumn("부품코드", width="medium"),
        "U/S": st.column_config.NumberColumn("U/S", min_value=0, default=1, width="small"),
        "재질/규격": st.column_config.TextColumn("재질/규격", width="medium"),
        "단위": st.column_config.TextColumn("단위", width="small"),
        "단가": st.column_config.NumberColumn("단가", min_value=0.0, format="%.1f", width="medium"),
        "NET(g,mm)": st.column_config.NumberColumn("NET(g,mm)", min_value=0.0, format="%.5f", width="medium"),
        "SCRAP(g,mm)": st.column_config.TextColumn("SCRAP(g,mm)", width="medium"),
        "자재LOSS율(%)": st.column_config.NumberColumn("자재LOSS율(%)", min_value=0.0, format="%.2f", width="medium"),
        "산업폐기물처리비용": st.column_config.NumberColumn("산업폐기물처리비용", min_value=0.0, format="%.2f", width="medium"),
        "다이캐스팅LOSS인정": st.column_config.NumberColumn("다이캐스팅LOSS인정", min_value=0.0, format="%.2f", width="medium"),
        "금액": st.column_config.NumberColumn("금액", format="%.2f", width="medium"),
        "LOSS금액": st.column_config.NumberColumn("LOSS금액", format="%.2f", width="medium"),
        "재료비": st.column_config.NumberColumn("재료비", format="%.2f", width="medium"),
    },
    key="material_editor",
    hide_index=True
)

# 세션 상태 업데이트 및 재계산
if not edited_mat.empty:
    # 입력 컬럼만 업데이트 (인덱스 리셋만 수행, 빈 행은 그대로 유지)
    input_cols = ['부품명', '부품코드', 'U/S', '재질/규격', '단위', '단가', 'NET(g,mm)', 'SCRAP(g,mm)', '자재LOSS율(%)', '산업폐기물처리비용', '다이캐스팅LOSS인정']
    updated_df = edited_mat[input_cols].copy().reset_index(drop=True)
    st.session_state.material_df = updated_df
    
    # 편집된 데이터로 재계산 (합계 표시용)
    final_calc = st.session_state.material_df.copy()
    final_calc['단가'] = pd.to_numeric(final_calc['단가'], errors='coerce').fillna(0)
    final_calc['U/S'] = pd.to_numeric(final_calc['U/S'], errors='coerce').fillna(1)
    final_calc['NET(g,mm)'] = pd.to_numeric(final_calc['NET(g,mm)'], errors='coerce').fillna(0)
    final_calc['자재LOSS율(%)'] = pd.to_numeric(final_calc['자재LOSS율(%)'], errors='coerce').fillna(0)
    final_calc['산업폐기물처리비용'] = pd.to_numeric(final_calc['산업폐기물처리비용'], errors='coerce').fillna(0)
    final_calc['다이캐스팅LOSS인정'] = pd.to_numeric(final_calc['다이캐스팅LOSS인정'], errors='coerce').fillna(0)

    final_calc['금액'] = final_calc['단가'] * final_calc['NET(g,mm)'] * final_calc['U/S']
    final_calc['LOSS금액'] = final_calc['금액'] * (final_calc['자재LOSS율(%)'] / 100)
    final_calc['재료비'] = final_calc['금액'] + final_calc['LOSS금액'] + final_calc['산업폐기물처리비용'] + final_calc['다이캐스팅LOSS인정']

    # 재료비 합계 계산 및 표시
    total_material_cost = final_calc['재료비'].sum()
    st.markdown("---")
    col1, col2 = st.columns([1, 3])
    with col1:
        st.metric("**재료비 합계**", f"₩ {total_material_cost:,.1f}")

st.divider()

# =========================================================
# [UI 5] 산출 결과 저장 / 불러오기
# =========================================================
st.header("💾 산출 결과 저장 / 불러오기")

# 1) 현재 산출 저장
col_save_left, col_save_right = st.columns([2, 3])
with col_save_left:
    save_name = st.text_input("저장 이름 (예: 96240-BQ000 1차 산출)", value=f"{p_no} - {p_name}")
    if st.button("📥 현재 산출 저장", type="primary", use_container_width=True):
        # 현재 재료비/가공비, 기본 정보 스냅샷
        snapshot = {
            "id": datetime.now().strftime("%Y%m%d%H%M%S"),
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "name": save_name,
            "p_no": p_no,
            "p_name": p_name,
            "car": car,
            "company": company,
            "labor_rate": labor_rate,
            "material": st.session_state.material_df.to_dict(orient="records") if "material_df" in st.session_state else [],
            "process": st.session_state.process_df.to_dict(orient="records") if "process_df" in st.session_state else [],
        }
        all_results = load_saved_results()
        all_results.append(snapshot)
        save_results_list(all_results)
        st.success("현재 산출이 저장되었습니다. 아래 목록에서 확인할 수 있습니다.")

# 2) 저장된 산출 목록
saved_results = load_saved_results()

st.subheader("📂 저장된 산출 목록")
if not saved_results:
    st.info("저장된 산출 결과가 없습니다. 먼저 위에서 산출을 저장해주세요.")
else:
    # 메타 정보용 테이블
    meta_rows = [
        {
            "번호": len(saved_results) - idx,
            "저장ID": item["id"],
            "저장일시": item.get("saved_at", ""),
            "품번": item.get("p_no", ""),
            "품명": item.get("p_name", ""),
            "차종": item.get("car", ""),
            "업체": item.get("company", ""),
            "이름": item.get("name", ""),
        }
        for idx, item in enumerate(reversed(saved_results))
    ]
    meta_df = pd.DataFrame(meta_rows)

    selected_id = st.selectbox(
        "불러올 산출 선택 (저장ID 기준)",
        options=[row["저장ID"] for row in meta_rows],
        format_func=lambda x: next((f"{row['저장ID']} - {row['이름']}" for row in meta_rows if row["저장ID"] == x), x),
    )

    st.dataframe(meta_df, use_container_width=True, hide_index=True)

    # 선택된 산출 상세 보기 / 불러오기
    if selected_id:
        target = next((item for item in saved_results if item["id"] == selected_id), None)
        if target:
            st.markdown("---")
            st.markdown("### 🔍 선택한 산출 상세")
            st.write(f"품번: {target.get('p_no', '')} / 품명: {target.get('p_name', '')} / 차종: {target.get('car', '')}")

            tab_mat, tab_pro = st.tabs(["재료비 데이터", "가공비 데이터"])
            with tab_mat:
                mat_df = pd.DataFrame(target.get("material", []))
                if not mat_df.empty:
                    st.dataframe(mat_df, use_container_width=True, hide_index=True)
                else:
                    st.info("저장된 재료비 데이터가 없습니다.")
            with tab_pro:
                pro_df = pd.DataFrame(target.get("process", []))
                if not pro_df.empty:
                    st.dataframe(pro_df, use_container_width=True, hide_index=True)
                else:
                    st.info("저장된 가공비 데이터가 없습니다.")

            col_load, col_note = st.columns([1, 2])
            with col_load:
                if st.button("↩️ 이 산출을 편집 화면으로 불러오기", use_container_width=True):
                    # 기본 정보 및 재료비/가공비를 현재 세션에 적용
                    st.session_state.material_df = pd.DataFrame(target.get("material", [])).reset_index(drop=True)
                    st.session_state.process_df = pd.DataFrame(target.get("process", [])).reset_index(drop=True) if target.get("process") else get_default_process_df()
                    st.success("선택한 산출의 재료비 데이터가 편집 테이블에 반영되었습니다.")
                    try:
                        st.rerun()
                    except AttributeError:
                        if hasattr(st, "experimental_rerun"):
                            st.experimental_rerun()

 # =========================================================
# [UI 3] 가공비 입력
# =========================================================
st.subheader("3. 가공비 명세서")
st.caption("가공비 정보를 입력하세요. 엑셀 양식과 동일하게 출력됩니다.")

# 세션 상태에 가공비 테이블이 없으면 초기화
if "process_df" not in st.session_state:
    st.session_state.process_df = get_default_process_df()

edited_pro = st.data_editor(
    st.session_state.process_df,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "부품명": st.column_config.TextColumn("부품명", width="medium"),
        "U/S": st.column_config.NumberColumn("U/S", min_value=0, default=1, width="small"),
        "공정명": st.column_config.TextColumn("공정명", width="medium"),
        "사용기계": st.column_config.TextColumn("사용기계", width="medium"),
        "인": st.column_config.NumberColumn("인", min_value=0, default=1, width="small"),
        "공수(SEC)": st.column_config.NumberColumn("공수(SEC)", min_value=0.0, format="%.1f", width="medium"),
        "준비시간(분)": st.column_config.NumberColumn("준비시간(분)", min_value=0.0, format="%.1f", width="medium"),
        "산출근거(원/HR)": st.column_config.NumberColumn("산출근거(원/HR)", min_value=0.0, format="%.0f", width="medium"),
        "여유율(%)": st.column_config.NumberColumn("여유율(%)", min_value=0.0, format="%.1f", width="medium"),
    },
    key="process_editor",
)

# 편집 결과를 세션에 반영
st.session_state.process_df = edited_pro.copy().reset_index(drop=True)

# 가공비 계산 및 표시
total_process_cost = 0.0
if not st.session_state.process_df.empty:
    # 숫자형 데이터 변환
    calc_pro = st.session_state.process_df.copy()
    calc_pro['U/S'] = pd.to_numeric(calc_pro['U/S'], errors='coerce').fillna(1)
    calc_pro['인'] = pd.to_numeric(calc_pro['인'], errors='coerce').fillna(1)
    calc_pro['공수(SEC)'] = pd.to_numeric(calc_pro['공수(SEC)'], errors='coerce').fillna(0)
    calc_pro['준비시간(분)'] = pd.to_numeric(calc_pro['준비시간(분)'], errors='coerce').fillna(0)
    calc_pro['산출근거(원/HR)'] = pd.to_numeric(calc_pro['산출근거(원/HR)'], errors='coerce').fillna(0)
    calc_pro['여유율(%)'] = pd.to_numeric(calc_pro.get('여유율(%)', 0), errors='coerce').fillna(0)
    
    # 가공비 계산: (공수(SEC) / 3600) × 임율 × 인 × U/S
    # 산출근거가 있으면 산출근거 사용, 없으면 적용임율 사용
    calc_pro['사용임율'] = calc_pro['산출근거(원/HR)'].apply(lambda x: x if x > 0 else labor_rate)
    calc_pro['가공비'] = (calc_pro['공수(SEC)'] / 3600) * calc_pro['사용임율'] * calc_pro['인'] * calc_pro['U/S']
    
    # 준비시간 가공비 계산 (분을 시간으로 변환)
    calc_pro['준비시간가공비'] = (calc_pro['준비시간(분)'] / 60) * calc_pro['사용임율'] * calc_pro['인'] * calc_pro['U/S']

    # 여유율 적용: 총가공비 = 가공비 × (1 + 여유율/100) + 준비시간가공비
    calc_pro['총가공비'] = calc_pro['가공비'] * (1 + calc_pro['여유율(%)'] / 100) + calc_pro['준비시간가공비']
    
    # 부품별 가공비 표시
    st.markdown("**부품별 가공비 산출**")
    display_cols = ['부품명', '공정명', '사용기계', '인', '공수(SEC)', '임율(원/HR)', '여유율(%)', '가공비', '준비시간(분)', '준비시간가공비', '총가공비']
    calc_pro['임율(원/HR)'] = calc_pro['사용임율']
    available_cols = [col for col in display_cols if col in calc_pro.columns]
    
    # 가공비가 0보다 큰 행만 표시
    display_df = calc_pro[calc_pro['총가공비'] > 0][available_cols].copy() if len(calc_pro[calc_pro['총가공비'] > 0]) > 0 else calc_pro[available_cols].copy()
    
    # 숫자 포맷팅
    for col in ['임율(원/HR)', '가공비', '준비시간가공비', '총가공비']:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "0.00")
    
    if '공수(SEC)' in display_df.columns:
        display_df['공수(SEC)'] = display_df['공수(SEC)'].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "0.0")
    if '준비시간(분)' in display_df.columns:
        display_df['준비시간(분)'] = display_df['준비시간(분)'].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "0.0")
    
    st.dataframe(display_df, use_container_width=True, hide_index=True)
    
    # 가공비 합계 계산 및 표시
    total_process_cost = calc_pro['총가공비'].sum()
    st.markdown("---")
    col1, col2 = st.columns([1, 3])
    with col1:
        st.metric("**가공비 합계**", f"{total_process_cost:,.2f} 원")

# =========================================================
# [UI 4] 실시간 미리보기
# =========================================================
st.markdown("---")
st.header("👀 미리보기")

# 재료비 계산 (편집된 재료비 데이터 기준)
try:
    preview_mat = st.session_state.material_df.copy()
except Exception:
    # 세션에 없으면 edited_mat 사용 (초기 로드 대응)
    preview_mat = edited_mat.copy()

if not preview_mat.empty:
    preview_mat['단가'] = pd.to_numeric(preview_mat.get('단가', 0), errors='coerce').fillna(0)
    preview_mat['U/S'] = pd.to_numeric(preview_mat.get('U/S', 1), errors='coerce').fillna(1)
    preview_mat['NET(g,mm)'] = pd.to_numeric(preview_mat.get('NET(g,mm)', 0), errors='coerce').fillna(0)
    # 금액 = 단가 × NET(g,mm) × U/S  (상단 산출 로직과 동일)
    preview_mat['예상금액'] = preview_mat['단가'] * preview_mat['NET(g,mm)'] * preview_mat['U/S']
else:
    preview_mat = pd.DataFrame(columns=['예상금액'])

# 합계
total_mat_cost = preview_mat['예상금액'].sum()
# 가공비 합계는 위에서 계산한 total_process_cost 사용 (실제 총가공비와 일치)
total_pro_cost = total_process_cost

# KPI 카드
m1, m2 = st.columns(2)
m1.metric("재료비 합계 (예상)", f"{total_mat_cost:,.0f} 원")
m2.metric("가공비 합계 (예상)", f"{total_pro_cost:,.0f} 원")

# =========================================================
# [엑셀 생성 및 다운로드]
# =========================================================
def generate_excel():
    try:
        wb = load_workbook("template.xlsx")

        # 시트 찾기
        target_sheet = None
        for s in wb.sheetnames:
            if "원가" in s or "견적" in s or "계산" in s:
                target_sheet = s
                break
        
        if target_sheet:
            ws = wb[target_sheet]
        else:
            ws = wb.active

        # -------------------------------------------------
        # [공통] 헤더 텍스트를 이용해 실제 컬럼 번호 찾기
        #        (템플릿에서 한 칸씩 밀리는 문제 방지용)
        # -------------------------------------------------
        def find_col(keyword: str, default_col: int) -> int:
            """시트 상단에서 헤더 텍스트를 찾아 실제 컬럼 번호를 반환"""
            try:
                for row in ws.iter_rows(min_row=1, max_row=40):
                    for cell in row:
                        if not cell.value:
                            continue
                        cell_str = str(cell.value)
                        if keyword in cell_str:
                            return cell.column
            except Exception:
                pass
            return default_col

        # 재료비 영역 실제 컬럼 위치
        mat_code_col = find_col("품번", COL_MAT_CODE)
        mat_name_col = find_col("부품명", COL_MAT_NAME)
        mat_us_col = find_col("U/S", COL_MAT_US)
        mat_spec_col = find_col("재질", COL_MAT_SPEC)
        mat_unit_col = find_col("단위", COL_MAT_UNIT)
        mat_price_col = find_col("단가", COL_MAT_PRICE)
        mat_net_col = find_col("NET", COL_MAT_NET)
        mat_scrap_col = find_col("SCRAP", COL_MAT_SCRAP)
        mat_input_col = find_col("투입", COL_MAT_INPUT)
        mat_lossrate_col = find_col("LOSS율", COL_MAT_LOSS_RATE)
        mat_waste_col = find_col("산업폐기물", COL_MAT_WASTE)
        mat_die_col = find_col("다이캐스팅", COL_MAT_DIE_LOSS)

        # -------------------------------------------------
        # [A] 파이썬 기준 가공비 총합 재계산 (화면과 동일 로직)
        # -------------------------------------------------
        total_process_cost_excel = 0.0
        try:
            process_df = st.session_state.get("process_df", pd.DataFrame())
            if not process_df.empty:
                calc_pro_excel = process_df.copy()
                calc_pro_excel['U/S'] = pd.to_numeric(calc_pro_excel['U/S'], errors='coerce').fillna(1)
                calc_pro_excel['인'] = pd.to_numeric(calc_pro_excel['인'], errors='coerce').fillna(1)
                calc_pro_excel['공수(SEC)'] = pd.to_numeric(calc_pro_excel['공수(SEC)'], errors='coerce').fillna(0)
                calc_pro_excel['준비시간(분)'] = pd.to_numeric(calc_pro_excel['준비시간(분)'], errors='coerce').fillna(0)
                calc_pro_excel['산출근거(원/HR)'] = pd.to_numeric(calc_pro_excel['산출근거(원/HR)'], errors='coerce').fillna(0)
                calc_pro_excel['여유율(%)'] = pd.to_numeric(calc_pro_excel.get('여유율(%)', 0), errors='coerce').fillna(0)

                # 사용임율: 산출근거가 있으면 그 값, 없으면 적용임율 사용
                calc_pro_excel['사용임율'] = calc_pro_excel['산출근거(원/HR)'].apply(
                    lambda x: x if x > 0 else labor_rate
                )
                # 기본 가공비
                calc_pro_excel['가공비'] = (
                    (calc_pro_excel['공수(SEC)'] / 3600) *
                    calc_pro_excel['사용임율'] *
                    calc_pro_excel['인'] *
                    calc_pro_excel['U/S']
                )
                # 준비시간 가공비
                calc_pro_excel['준비시간가공비'] = (
                    (calc_pro_excel['준비시간(분)'] / 60) *
                    calc_pro_excel['사용임율'] *
                    calc_pro_excel['인'] *
                    calc_pro_excel['U/S']
                )
                # 여유율 적용 총가공비
                calc_pro_excel['총가공비'] = (
                    calc_pro_excel['가공비'] * (1 + calc_pro_excel['여유율(%)'] / 100) +
                    calc_pro_excel['준비시간가공비']
                )
                total_process_cost_excel = float(calc_pro_excel['총가공비'].sum())
        except Exception:
            total_process_cost_excel = 0.0

        # 1. 기본 정보 입력
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value:
                    cell_str = str(cell.value)
                    # 품번
                    if "품 번" in cell_str or "품번" in cell_str:
                        safe_write(ws, cell.row, cell.column+2, p_no)
                    # 품명
                    if "품명" in cell_str and "부품명" not in cell_str:
                        safe_write(ws, cell.row, cell.column+2, p_name)
                    # 차종
                    if "차종" in cell_str:
                        safe_write(ws, cell.row, cell.column+1, car)
                    # 업체
                    if "업체" in cell_str:
                        safe_write(ws, cell.row, cell.column+1, company)
                    # 적용임율
                    if "적용임율" in cell_str or ("임율" in cell_str and "적용" in cell_str):
                        safe_write(ws, cell.row, cell.column+1, labor_rate)

        # 2. 재료비 입력 (기존 데이터 지우기)
        for r in range(MAT_START_ROW, MAT_MAX_ROW + 1):
            safe_write(ws, r, mat_code_col, "")
            safe_write(ws, r, mat_name_col, "")
            safe_write(ws, r, mat_us_col, None)
            safe_write(ws, r, mat_spec_col, "")
            safe_write(ws, r, mat_unit_col, "")
            safe_write(ws, r, mat_price_col, None)
            safe_write(ws, r, mat_net_col, None)
            safe_write(ws, r, mat_scrap_col, None)
            safe_write(ws, r, mat_input_col, None)
            safe_write(ws, r, mat_lossrate_col, None)
            safe_write(ws, r, mat_waste_col, None)
            safe_write(ws, r, mat_die_col, None)

        # 재료비 데이터 쓰기
        current_row = MAT_START_ROW
        for idx, row in edited_mat.iterrows():
            if current_row > MAT_MAX_ROW:
                break
            if pd.notna(row.get('부품명')) and str(row.get('부품명', '')).strip():
                # 품번 / 부품명
                safe_write(ws, current_row, mat_code_col, row.get('부품코드', ''))
                safe_write(ws, current_row, mat_name_col, row.get('부품명', ''))
                # U/S 는 반드시 숫자 (화면의 U/S 값)
                safe_write(ws, current_row, mat_us_col, row.get('U/S', 1))
                # 재질/규격에는 지금까지 U/S 열에 들어가던 정보를 넣어야 한다고 요청하셨음
                # 현재 화면 구조상 이 값은 별도 컬럼 '재질/규격' 에 들어있으므로 우선 그 값을 사용
                # (필요 시 부품코드 등을 추가로 입력 가능)
                safe_write(ws, current_row, mat_spec_col, row.get('재질/규격', ''))
                # 나머지 단위/단가/NET 등
                safe_write(ws, current_row, mat_unit_col, row.get('단위', 'EA'))
                safe_write(ws, current_row, mat_price_col, row.get('단가', 0))
                safe_write(ws, current_row, mat_net_col, row.get('NET(g,mm)', 0))
                # SCRAP은 텍스트이므로 숫자 변환 시도
                scrap_val = row.get('SCRAP(g,mm)', '')
                scrap_num = pd.to_numeric(scrap_val, errors='coerce')
                safe_write(ws, current_row, mat_scrap_col, scrap_num if pd.notna(scrap_num) else None)
                # 투입중량은 NET(g,mm)와 동일하게 설정 (엑셀 양식에 따라)
                safe_write(ws, current_row, mat_input_col, row.get('NET(g,mm)', 0))
                safe_write(ws, current_row, mat_lossrate_col, row.get('자재LOSS율(%)', 0))
                safe_write(ws, current_row, mat_waste_col, row.get('산업폐기물처리비용', 0))
                safe_write(ws, current_row, mat_die_col, row.get('다이캐스팅LOSS인정', 0))
                current_row += 1

        # 3. 가공비 입력 (기존 데이터 지우기)
        for r in range(PRO_START_ROW, PRO_MAX_ROW + 1):
            safe_write(ws, r, COL_PRO_NAME, "")
            safe_write(ws, r, COL_PRO_US, None)
            safe_write(ws, r, COL_PRO_PROCESS, "")
            safe_write(ws, r, COL_PRO_MACH, "")
            safe_write(ws, r, COL_PRO_MAN, None)
            safe_write(ws, r, COL_PRO_TIME, None)
            safe_write(ws, r, COL_PRO_PREP, None)
            safe_write(ws, r, COL_PRO_BASIS, None)
            # 임율은 기본값으로 설정
            safe_write(ws, r, COL_PRO_RATE, labor_rate)

        # 3. 가공비 데이터 쓰기 (행별 금액은 화면에서 계산한 총가공비 사용)
        process_df = st.session_state.get("process_df", pd.DataFrame())
        if not process_df.empty:
            # 시트에서 공정명 텍스트를 기준으로 실제 행 번호를 매핑
            process_row_map: dict[str, int] = {}
            try:
                for r in range(PRO_START_ROW, PRO_MAX_ROW + 20):
                    cell = ws.cell(row=r, column=COL_PRO_PROCESS)
                    if cell.value:
                        key = str(cell.value).strip()
                        process_row_map[key] = r
            except Exception:
                process_row_map = {}

            # 위에서 계산한 calc_pro_excel과 인덱스를 맞춰 사용
            calc_table = calc_pro_excel if 'calc_pro_excel' in locals() else process_df

            for idx, row in process_df.iterrows():
                proc_name = str(row.get('공정명', '')).strip()
                if not proc_name:
                    continue

                # 템플릿 상에서 동일한 공정명을 가진 행을 찾아서 그 위치에 써준다
                target_row = process_row_map.get(proc_name)
                if not target_row:
                    # 못 찾으면 해외 가공비 표 안에서 순차 배치 (fallback)
                    for r in range(PRO_START_ROW, PRO_MAX_ROW + 1):
                        if not ws.cell(row=r, column=COL_PRO_PROCESS).value:
                            target_row = r
                            break
                if not target_row:
                    continue

                safe_write(ws, target_row, COL_PRO_NAME, row.get('부품명', ''))
                safe_write(ws, target_row, COL_PRO_US, row.get('U/S', 1))
                safe_write(ws, target_row, COL_PRO_PROCESS, proc_name)
                safe_write(ws, target_row, COL_PRO_MACH, row.get('사용기계', ''))
                safe_write(ws, target_row, COL_PRO_MAN, row.get('인', 1))
                safe_write(ws, target_row, COL_PRO_TIME, row.get('공수(SEC)', 0))
                safe_write(ws, target_row, COL_PRO_PREP, row.get('준비시간(분)', 0))
                safe_write(ws, target_row, COL_PRO_BASIS, row.get('산출근거(원/HR)', 0))

                # 사용임율(임율/HR)
                use_rate = float(calc_table.loc[idx, '사용임율']) if '사용임율' in calc_table.columns else float(labor_rate)
                safe_write(ws, target_row, COL_PRO_RATE, use_rate)

                # 행별 금액(원/EA) = 총가공비 (화면과 동일)
                row_total = float(calc_table.loc[idx, '총가공비']) if '총가공비' in calc_table.columns else 0.0
                safe_write(ws, target_row, COL_PRO_AMOUNT1, row_total)

            # -------------------------------------------------
            # [특수 처리] 국내 가공비 두 행을 템플릿 고정 위치에 강제로 반영
            #  - 하역/리패킹/검사  → 44행
            #  - 라벨/포장/출하     → 45행
            # (화면 값은 맞는데 엑셀에서 해당 행 금액이 표시되지 않는 문제 보완)
            # -------------------------------------------------
            try:
                DOM_START_ROW = 44
                domestic_names = ["하역/리패킹/검사", "라벨/포장/출하"]
                for offset, dname in enumerate(domestic_names):
                    mask = process_df.get("공정명", "").astype(str).str.contains(dname)
                    if not mask.any():
                        continue
                    idx = process_df[mask].index[0]
                    row = process_df.loc[idx]
                    target_row = DOM_START_ROW + offset

                    safe_write(ws, target_row, COL_PRO_NAME, row.get('부품명', ''))
                    safe_write(ws, target_row, COL_PRO_US, row.get('U/S', 1))
                    safe_write(ws, target_row, COL_PRO_PROCESS, str(row.get('공정명', '')))
                    safe_write(ws, target_row, COL_PRO_MACH, row.get('사용기계', ''))
                    safe_write(ws, target_row, COL_PRO_MAN, row.get('인', 1))
                    safe_write(ws, target_row, COL_PRO_TIME, row.get('공수(SEC)', 0))
                    safe_write(ws, target_row, COL_PRO_PREP, row.get('준비시간(분)', 0))
                    safe_write(ws, target_row, COL_PRO_BASIS, row.get('산출근거(원/HR)', 0))

                    use_rate = float(calc_table.loc[idx, '사용임율']) if '사용임율' in calc_table.columns else float(labor_rate)
                    safe_write(ws, target_row, COL_PRO_RATE, use_rate)

                    row_total = float(calc_table.loc[idx, '총가공비']) if '총가공비' in calc_table.columns else 0.0
                    safe_write(ws, target_row, COL_PRO_AMOUNT1, row_total)
            except Exception:
                # 실패해도 전체 생성에는 영향 없게 처리
                pass

        # 5. 엑셀 하단 (4)가공비 합계를 파이썬에서 계산한 total_process_cost_excel로 덮어쓰기
        #    템플릿 수식과 파이썬 로직이 다를 수 있으므로, 최종 합계만은 일치시키기 위함
        try:
            found_gagongbi_row = False
            for row in ws.iter_rows(min_row=1, max_row=200):
                for cell in row:
                    if not cell.value:
                        continue
                    cell_str = str(cell.value).replace(" ", "")
                    # "(4)가공비" 또는 "가공비"가 포함된 셀을 찾는다
                    if "가공비" in cell_str and "(4)" in cell_str:
                        # 같은 행에서 가장 오른쪽에 있는 숫자/수식 셀을 찾아서 거기에 총가공비를 쓴다
                        target_cell = None
                        for c2 in row:
                            if c2.column <= cell.column:
                                continue
                            if isinstance(c2.value, (int, float)) or (
                                isinstance(c2.value, str) and c2.value.startswith("=")
                            ):
                                target_cell = c2
                        if target_cell is None:
                            # 못 찾으면 오른쪽 몇 칸 뒤에 강제로 기록
                            target_cell = ws.cell(row=cell.row, column=cell.column + 5)
                        target_cell.value = total_process_cost_excel
                        found_gagongbi_row = True
                        break
                if found_gagongbi_row:
                    break
        except Exception:
            # 실패해도 엑셀 저장 자체는 계속 진행
            pass

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except FileNotFoundError:
        return f"ERROR: template.xlsx 파일을 찾을 수 없습니다.\n프로젝트 폴더에 template.xlsx 파일이 있는지 확인해주세요."
    except Exception as e:
        return f"ERROR: {str(e)}\n\n{traceback.format_exc()}"

# 다운로드 버튼
st.markdown("---")
if st.button("✅ 엑셀 파일 생성 및 다운로드", type="primary", use_container_width=True):
    result = generate_excel()
    
    if isinstance(result, str) and result.startswith("ERROR"):
        st.error("오류가 발생했습니다.")
        st.text(result)
    else:
        st.success("엑셀 파일이 생성되었습니다!")
        st.download_button(
            label="📥 원가계산서 다운로드",
            data=result,
            file_name=f"원가계산서_{p_no}_{p_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
